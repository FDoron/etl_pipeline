import pandas as pd
import os
from src.utils.logger import logger
from src.ingestion.file_ops import ensure_filename_suffix, move_file
from src.transformation.transform import normalize_and_validate
from src.db.db_ops import insert_dataframe, ProcessingJob
from datetime import datetime
from sqlalchemy.exc import OperationalError
from src.transformation.data_prep import prepare_data
from charset_normalizer import detect
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from src.utils.config_loader import Config

Config.load()


def ingest_file(file_path, provider_mapping, session, job_id):
    try:
        # Ensure filename has date suffix
        file_path = ensure_filename_suffix(file_path)

        # Read file with encoding
        try:
            # In ingest_file, replace CSV reading block (line ~30-46)
            if file_path.endswith('.csv'):
                with open(file_path, 'rb') as f:
                    result = detect(f.read(1024))
                encoding = result.get('encoding', 'utf-8')
                logger.debug(f"Detected encoding for {file_path}: {encoding}", extra={"job_id": job_id})
                try:
                    # Read first line to check for headers
                    with open(file_path, 'r', encoding=encoding) as f:
                        first_line = f.readline().strip().split(',')
                    # Assume headers if first line contains non-numeric strings matching expected columns
                    expected_columns = ['ID', 'fee', 'provider', 'paid_month']  # From settings.yaml
                    has_headers = any(col.lower() in [x.lower() for x in first_line] for col in expected_columns)
                    try:
                        df = pd.read_csv(file_path, encoding=encoding, header=0 if has_headers else None)
                        if not has_headers:
                            num_cols = len(df.columns)
                            if num_cols < len(expected_columns):
                                logger.error(f"CSV has fewer columns ({num_cols}) than required ({len(expected_columns)})", extra={"job_id": job_id, "file": file_path})
                                session.query(ProcessingJob).filter_by(job_id=job_id).update({
                                    'status': 'FAILED',
                                    'error_summary': f'CSV has insufficient columns: {num_cols}',
                                    'finished_at': datetime.utcnow()
                                })
                                session.commit()
                                move_file(file_path, 'data/failed', suffix='failed')
                                return False
                            # Assign col_{running_number} for all columns
                            df.columns = [f'col_{i}' for i in range(num_cols)]
                            logger.warning(f"No headers detected in {file_path}, assigned columns: {df.columns.tolist()}", extra={"job_id": job_id})
                    except UnicodeDecodeError:
                        df = pd.read_csv(file_path, encoding='cp1255', header=0 if has_headers else None)
                        if not has_headers:
                            num_cols = len(df.columns)
                            if num_cols < len(expected_columns):
                                logger.error(f"CSV has fewer columns ({num_cols}) than required ({len(expected_columns)})", extra={"job_id": job_id, "file": file_path})
                                session.query(ProcessingJob).filter_by(job_id=job_id).update({
                                    'status': 'FAILED',
                                    'error_summary': f'CSV has insufficient columns: {num_cols}',
                                    'finished_at': datetime.utcnow()
                                })
                                session.commit()
                                move_file(file_path, 'data/failed', suffix='failed')
                                return False
                            df.columns = [f'col_{i}' for i in range(num_cols)]
                            logger.warning(f"No headers detected in {file_path}, assigned columns: {df.columns.tolist()}", extra={"job_id": job_id})
                except Exception as e:
                    logger.error(f"Failed to read file", extra={"job_id": job_id, "file": file_path, "error": str(e)})
                    session.query(ProcessingJob).filter_by(job_id=job_id).update({
                        'status': 'FAILED',
                        'error_summary': f'Failed to read file: {str(e)}',
                        'finished_at': datetime.utcnow()
                    })
                    session.commit()
                    move_file(file_path, 'data/failed', suffix='failed')
                    return False
            elif file_path.endswith(('.xls', '.xlsx')):
                df = pd.read_excel(file_path)
                text = ' '.join(df.iloc[:, :].astype(str).values.flatten()).encode()
                result = detect(text)
                encoding = result.get('encoding', 'utf-8')
                logger.debug(f"Detected encoding for {file_path}: {encoding}", extra={"job_id": job_id})
            elif file_path.endswith('.txt'):
                df = pd.read_table(file_path)
            else:
                logger.error("Unsupported file format", extra={"job_id": job_id, "file": file_path})
                session.query(ProcessingJob).filter_by(job_id=job_id).update({
                    'status': 'FAILED',
                    'error_summary': 'Unsupported file format',
                    'finished_at': datetime.utcnow()
                })
                session.commit()
                move_file(file_path, 'data/failed', suffix='failed')
                return False
        except Exception as e:
            logger.error(f"Failed to read file", extra={"job_id": job_id, "file": file_path, "error": str(e)})
            session.query(ProcessingJob).filter_by(job_id=job_id).update({
                'status': 'FAILED',
                'error_summary': f'Failed to read file: {str(e)}',
                'finished_at': datetime.utcnow()
            })
            session.commit()
            move_file(file_path, 'data/failed', suffix='failed')
            return False

        provider_mapping['file_path'] = file_path
        original_df = df.copy()  # Store original for comparison
        result = prepare_data(df, provider_mapping, session, job_id)
        if result['status'] == 'failed':
            logger.error(f"Data preparation failed for {file_path}", extra={"job_id": job_id, "reason": result.get('reason', 'Unknown error')})
            session.query(ProcessingJob).filter_by(job_id=job_id).update({
                'status': 'FAILED',
                'error_summary': result.get('reason', 'Unknown error'),
                'finished_at': datetime.utcnow()
            })
            session.commit()
            move_file(file_path, 'data/failed', suffix='failed')
            return False

        valid_df, result_dict = normalize_and_validate(result['df'], {'column_mapping': provider_mapping['column_mapping'], 'provider': provider_mapping['provider']}, session, job_id)
        invalid_rows = result_dict.get('invalid_rows', [])
        if invalid_rows:
            logger.warning(f"Found {len(invalid_rows)} invalid rows for job_id={job_id}", extra={"job_id": job_id, "invalid_rows": invalid_rows})
        
        if result_dict['status'] == 'failed':
            logger.error(f"Validation failed for {file_path}", extra={"job_id": job_id, "reason": result_dict['reason']})
            session.query(ProcessingJob).filter_by(job_id=job_id).update({
                'status': 'FAILED',
                'error_summary': result_dict['reason'],
                'finished_at': datetime.utcnow()
            })
            session.commit()
            move_file(file_path, 'data/archive', suffix='failed')  # Archive for showstopper
            return False

        # Update job status
        rows_processed = len(original_df)
        rows_failed = len(invalid_rows)
        rows_inserted = len(valid_df) if valid_df is not None else 0

        try:
            session.query(ProcessingJob).filter_by(job_id=job_id).update({
                'rows_processed': rows_processed,
                'rows_failed': rows_failed,
                'rows_inserted': rows_inserted,
                'status': 'PARTIAL' if rows_failed > 0 and rows_inserted > 0 else ('SUCCESS' if rows_inserted > 0 else 'FAILED'),
                'finished_at': datetime.utcnow()
            }, synchronize_session='evaluate')
            logger.debug(f"Updated ProcessingJob for job_id={job_id}, rows_processed={rows_processed}", extra={"job_id": job_id})
            session.commit()
        except OperationalError as e:
            logger.error(f"Database update failed for job_id={job_id}", extra={"job_id": job_id, "error": str(e)})
            session.rollback()
            move_file(file_path, 'data/archive', suffix='failed')
            return False

        if valid_df is None:
            logger.error(f"No valid rows after validation", extra={"job_id": job_id, "file": file_path})
            session.query(ProcessingJob).filter_by(job_id=job_id).update({
                'status': 'FAILED',
                'error_summary': 'No valid rows after validation',
                'finished_at': datetime.utcnow()
            })
            session.commit()
            move_file(file_path, 'data/archive', suffix='failed')
            return False


        # In ingest_file, before insert_dataframe (around line ~170)
        column_mapping = {
            'ID': 'customer_id',
            'fee': 'monthly_fee',
            'provider': 'provider_name',
            'paid_month': 'paid_month',
            'ingested_at': 'ingested_at',
            'job_id': 'job_id',
            'status': 'status'
        }
        valid_df = valid_df.rename(columns={k: v for k, v in column_mapping.items() if k in valid_df.columns})
        logger.debug(f"Mapped columns for job_id={job_id}: {list(valid_df.columns)}", extra={"job_id": job_id})
        # Insert valid data
        # Fetch ProcessingJob to get file_name, report_period, job_id, status
        job = session.query(ProcessingJob).filter_by(job_id=job_id).first()
        original_name = os.path.splitext(os.path.basename(job.file_name))[0]  # Extract original name without extension
        reported_date = job.report_period.replace('-', '')  # Convert e.g., '09-2025' to '092025'
        status = job.status  # Use ProcessingJob status (STARTED, SUCCESS, FAILED, PARTIAL)

        # Save review file (invalid rows) with naming: original_name_reported_date_job_id_status_review.xlsx
        if invalid_rows:
            invalid_df = pd.DataFrame([row['data'] for row in invalid_rows])
            invalid_df['row_index'] = [row['row_index'] for row in invalid_rows]
            invalid_df['errors'] = [', '.join(row['errors']) for row in invalid_rows]
            review_file = f"data/review/{original_name}_{reported_date}_{job_id}_{status}_review.xlsx"
            invalid_df.to_excel(review_file, index=False)
            logger.info(f"Saved {len(invalid_rows)} invalid rows to {review_file}", extra={"job_id": job_id})

        # Save highlighted file (all rows, invalid highlighted) with naming: original_name_reported_date_job_id_status_highlighted.xlsx
        valid_df['errors'] = ''
        if invalid_rows:
            invalid_df['errors'] = invalid_df['errors'].apply(lambda x: '; '.join(x))
            combined_df = pd.concat([valid_df, invalid_df]).sort_index()
        else:
            combined_df = valid_df
        highlighted_file = f"data/review/{original_name}_{reported_date}_{job_id}_{status}_highlighted.xlsx"
        combined_df.to_excel(highlighted_file, index=False)
        wb = load_workbook(highlighted_file)
        ws = wb.active
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        invalid_indices = [row['row_index'] for row in invalid_rows]
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
            if combined_df.index[idx-1] in invalid_indices:
                for cell in row:
                    cell.fill = red_fill
        wb.save(highlighted_file)
        logger.info(f"Saved highlighted original to {highlighted_file}", extra={"job_id": job_id})

        # Move input file to data/review with naming: original_name_reported_date_job_id_status.extension
        new_file_path = move_file(job.file_name, 'data/review', job_id, status, job.report_period)
        if not new_file_path:
            logger.error(f"Failed to move input file for job_id={job_id}", extra={"job_id": job_id})
            session.query(ProcessingJob).filter_by(job_id=job_id).update({
                'status': 'FAILED',
                'error_summary': 'Failed to move input file',
                'finished_at': datetime.utcnow()
            })
            session.commit()
            return False

        # Update ProcessingJob with new file path
        session.query(ProcessingJob).filter_by(job_id=job_id).update({
            'file_name': new_file_path,
            'rows_processed': len(original_df),
            'rows_failed': len(invalid_rows),
            'rows_inserted': len(valid_df) if not Config.get('ingestion.skip_reports_insert', False) else 0,
            'finished_at': datetime.utcnow(),
            'status': 'SUCCESS' if not invalid_rows else 'PARTIAL'
        })
        session.commit()
        logger.info(f"Completed ingestion for job_id={job_id} with {len(invalid_rows)} invalid rows", extra={"job_id": job_id})
        return True

    # Replace existing exception block
    except Exception as e:
        logger.error(f"Ingestion failed", extra={"job_id": job_id, "file": file_path, "error": str(e)})
        session.query(ProcessingJob).filter_by(job_id=job_id).update({
            'status': 'FAILED',
            'error_summary': str(e),
            'finished_at': datetime.utcnow()
        })
        session.commit()
        move_file(file_path, 'data/failed', job_id, 'FAILED', job.report_period if job else 'unknown')
        return False